"""Fail-closed listed-company identity resolution backed by official snapshots."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import UTC, datetime
from difflib import SequenceMatcher
from enum import Enum
from functools import lru_cache
import io
import json
import os
from pathlib import Path
import re
from typing import Any, Callable, Iterable
import unicodedata
from uuid import uuid4

import requests


SECURITY_IDENTITY_SCHEMA_VERSION = "1.0"
SECURITY_MARKETS = ("CN", "HK", "US")

CNINFO_STOCK_URL = "https://www.cninfo.com.cn/new/data/szse_stock.json"
HKEX_ACTIVE_STOCK_URL = (
    "https://www1.hkexnews.hk/ncms/script/eds/activestock_sehk_c.json"
)
HKEX_INACTIVE_STOCK_URL = (
    "https://www1.hkexnews.hk/ncms/script/eds/inactivestock_sehk_c.json"
)
HKEX_SECURITIES_URL = (
    "https://www.hkex.com.hk/eng/services/trading/securities/"
    "securitieslists/ListOfSecurities.xlsx"
)
HKEX_STANDARD_TRANSFER_URL = (
    "https://www.hkex.com.hk/-/media/HKEX-Market/Services/Trading/Securities/"
    "Securities-Lists/Securities-Using-Standard-Transfer-Form-%28including-GEM%29-"
    "By-Stock-Code-Order/secstkorder.xls"
)
SEC_TICKER_URL = "https://www.sec.gov/files/company_tickers.json"
NASDAQ_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
NASDAQ_OTHER_LISTED_URL = (
    "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"
)

_MARKET_ORDER = {market: index for index, market in enumerate(SECURITY_MARKETS)}
_FUZZY_CANDIDATE_THRESHOLD = 0.68
_FUZZY_RESOLVE_THRESHOLD = 0.92
_FUZZY_MARGIN = 0.08
_DEFAULT_MINIMUM_RECORDS = {"CN": 1_000, "HK": 1_000, "US": 1_000}


class SecurityIdentityError(ValueError):
    """Raised when security-master data violates the identity contract."""


class SecurityMasterUnavailableError(SecurityIdentityError):
    """Raised when no validated local security-master snapshot is available."""


class SecurityIdentityResolutionError(SecurityIdentityError):
    """Raised when a caller tries to continue without one resolved identity."""

    def __init__(self, result: "IdentityResult"):
        self.result = result
        super().__init__(json.dumps(result.to_dict(), ensure_ascii=False, sort_keys=True))


class IdentityStatus(str, Enum):
    RESOLVED = "resolved"
    AMBIGUOUS = "ambiguous"
    CONFLICT = "conflict"
    MISSING = "missing"


def _required_text(value: Any, name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SecurityIdentityError(f"{name} must be non-empty text")
    return value.strip()


def _market(value: Any) -> str:
    market = _required_text(value, "market").upper()
    if market not in SECURITY_MARKETS:
        raise SecurityIdentityError(f"unsupported market: {market}")
    return market


def _normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return "".join(character for character in normalized if character.isalnum())


@lru_cache(maxsize=1)
def _hk_t2s_converter() -> Any:
    try:
        import opencc
    except ImportError as exc:
        raise SecurityIdentityError(
            "HK security-master refresh requires the catalog extra (opencc)"
        ) from exc
    return opencc.OpenCC("t2s.json")


@lru_cache(maxsize=8_192)
def _hk_simplified(value: str) -> str:
    return str(_hk_t2s_converter().convert(value)).strip()


def _ticker_key(value: str, market: str) -> str:
    text = unicodedata.normalize("NFKC", value).strip().upper()
    if market == "CN":
        text = re.sub(r"^(?:SH|SZ|BJ)[.:\-]?", "", text)
        text = re.sub(r"[.:\-]?(?:SH|SZ|BJ)$", "", text)
        return text.zfill(6) if text.isdigit() else text
    if market == "HK":
        text = re.sub(r"[.:\-]?HK$", "", text)
        return text.zfill(5) if text.isdigit() and len(text) <= 5 else text
    return re.sub(r"[^A-Z0-9]", "", text)


def _generated_aliases(name: str, market: str) -> tuple[str, ...]:
    values = {name, unicodedata.normalize("NFKC", name).strip()}
    if market in {"CN", "HK"}:
        values.update(
            re.sub(r"-(?:W|SW|B)$", "", value, flags=re.IGNORECASE)
            for value in tuple(values)
        )
        if market == "HK":
            values.update(_hk_simplified(value) for value in tuple(values))
        suffixes = (
            "股份有限公司",
            "有限责任公司",
            "有限責任公司",
            "有限公司",
            "集团",
            "集團",
            "控股",
            "公司",
        )
        for value in tuple(values):
            for suffix in suffixes:
                if value.endswith(suffix):
                    values.add(value[: -len(suffix)])
    else:
        words = re.split(r"\s+", re.sub(r"[^A-Za-z0-9]+", " ", name).strip())
        corporate = {
            "company",
            "co",
            "corporation",
            "corp",
            "incorporated",
            "inc",
            "limited",
            "ltd",
            "plc",
            "holdings",
            "holding",
        }
        while words and words[-1].casefold() in corporate:
            words.pop()
        if words:
            values.add(" ".join(words))
    return tuple(
        sorted(
            value.strip()
            for value in values
            if value.strip() and len(_normalize_text(value)) >= 2 and value.strip() != name
        )
    )


@dataclass(frozen=True)
class SecurityRecord:
    canonical_name: str
    market: str
    exchange: str
    ticker: str
    security_id: str
    aliases: tuple[str, ...]
    active: bool
    source_name: str
    source_url: str
    source_record_id: str
    identifiers: dict[str, str]
    schema_version: str = SECURITY_IDENTITY_SCHEMA_VERSION

    def __post_init__(self) -> None:
        if self.schema_version != SECURITY_IDENTITY_SCHEMA_VERSION:
            raise SecurityIdentityError(
                f"schema_version must be {SECURITY_IDENTITY_SCHEMA_VERSION}"
            )
        object.__setattr__(
            self, "canonical_name", _required_text(self.canonical_name, "canonical_name")
        )
        object.__setattr__(self, "market", _market(self.market))
        object.__setattr__(self, "exchange", _required_text(self.exchange, "exchange").upper())
        object.__setattr__(self, "ticker", _required_text(self.ticker, "ticker").upper())
        object.__setattr__(
            self, "security_id", _required_text(self.security_id, "security_id").upper()
        )
        if not isinstance(self.active, bool):
            raise SecurityIdentityError("active must be boolean")
        aliases = tuple(
            sorted(
                {
                    _required_text(alias, "alias")
                    for alias in self.aliases
                    if _normalize_text(str(alias))
                },
                key=lambda value: (_normalize_text(value), value),
            )
        )
        object.__setattr__(self, "aliases", aliases)
        object.__setattr__(self, "source_name", _required_text(self.source_name, "source_name"))
        source_url = _required_text(self.source_url, "source_url")
        if not source_url.startswith("https://"):
            raise SecurityIdentityError("source_url must use https")
        object.__setattr__(self, "source_url", source_url)
        object.__setattr__(
            self,
            "source_record_id",
            _required_text(self.source_record_id, "source_record_id"),
        )
        if not isinstance(self.identifiers, dict) or any(
            not isinstance(key, str) or not isinstance(value, str)
            for key, value in self.identifiers.items()
        ):
            raise SecurityIdentityError("identifiers must be a string mapping")
        object.__setattr__(self, "identifiers", dict(sorted(self.identifiers.items())))

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "canonical_name": self.canonical_name,
            "market": self.market,
            "exchange": self.exchange,
            "ticker": self.ticker,
            "security_id": self.security_id,
            "aliases": list(self.aliases),
            "active": self.active,
            "source_name": self.source_name,
            "source_url": self.source_url,
            "source_record_id": self.source_record_id,
            "identifiers": self.identifiers,
        }

    @classmethod
    def from_dict(cls, value: Any) -> "SecurityRecord":
        if not isinstance(value, dict):
            raise SecurityIdentityError("security record must be an object")
        expected = {
            "schema_version",
            "canonical_name",
            "market",
            "exchange",
            "ticker",
            "security_id",
            "aliases",
            "active",
            "source_name",
            "source_url",
            "source_record_id",
            "identifiers",
        }
        if set(value) != expected:
            raise SecurityIdentityError("security record fields do not match schema")
        aliases = value["aliases"]
        if not isinstance(aliases, list):
            raise SecurityIdentityError("aliases must be a list")
        return cls(
            schema_version=value["schema_version"],
            canonical_name=value["canonical_name"],
            market=value["market"],
            exchange=value["exchange"],
            ticker=value["ticker"],
            security_id=value["security_id"],
            aliases=tuple(aliases),
            active=value["active"],
            source_name=value["source_name"],
            source_url=value["source_url"],
            source_record_id=value["source_record_id"],
            identifiers=value["identifiers"],
        )


@dataclass(frozen=True)
class SecurityMaster:
    records: tuple[SecurityRecord, ...]
    markets: tuple[str, ...]
    snapshots: dict[str, dict[str, Any]]


def _record_sort_key(record: SecurityRecord) -> tuple[Any, ...]:
    return (
        _MARKET_ORDER[record.market],
        record.exchange,
        _ticker_key(record.ticker, record.market),
        record.canonical_name,
    )


class SecurityMasterStore:
    """Versioned per-market JSON snapshots with atomic replacement."""

    def __init__(self, cache_dir: Path):
        if not isinstance(cache_dir, Path):
            raise TypeError("cache_dir must be pathlib.Path")
        self.cache_dir = cache_dir.resolve(strict=False)

    def path_for(self, market: str) -> Path:
        return self.cache_dir / f"{_market(market).lower()}.json"

    def has_market(self, market: str) -> bool:
        return self.path_for(market).is_file()

    def write_market(
        self,
        market: str,
        records: Iterable[SecurityRecord],
        *,
        retrieved_at: str,
        sources: tuple[str, ...],
    ) -> Path:
        selected_market = _market(market)
        materialized = tuple(sorted(records, key=_record_sort_key))
        if any(record.market != selected_market for record in materialized):
            raise SecurityIdentityError("snapshot contains a record from another market")
        identities = [(record.market, record.security_id) for record in materialized]
        if len(identities) != len(set(identities)):
            raise SecurityIdentityError("snapshot contains duplicate market/security_id records")
        retrieved_at = _required_text(retrieved_at, "retrieved_at")
        source_values = tuple(sorted({_required_text(item, "source") for item in sources}))
        payload = {
            "schema_version": SECURITY_IDENTITY_SCHEMA_VERSION,
            "market": selected_market,
            "retrieved_at": retrieved_at,
            "sources": list(source_values),
            "record_count": len(materialized),
            "records": [record.to_dict() for record in materialized],
        }
        path = self.path_for(selected_market)
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_name(f"{path.name}.{os.getpid()}.{uuid4().hex}.tmp")
        try:
            temporary.write_text(
                json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n",
                encoding="utf-8",
                newline="\n",
            )
            os.replace(temporary, path)
        finally:
            temporary.unlink(missing_ok=True)
        return path

    def load(
        self,
        *,
        markets: tuple[str, ...] | None = None,
        require_all: bool = False,
    ) -> SecurityMaster:
        requested = tuple(_market(item) for item in (markets or SECURITY_MARKETS))
        records: list[SecurityRecord] = []
        snapshots: dict[str, dict[str, Any]] = {}
        loaded_markets: list[str] = []
        for market in requested:
            path = self.path_for(market)
            if not path.is_file():
                continue
            try:
                payload = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise SecurityIdentityError(f"invalid {market} security snapshot: {exc}") from exc
            if not isinstance(payload, dict) or set(payload) != {
                "schema_version",
                "market",
                "retrieved_at",
                "sources",
                "record_count",
                "records",
            }:
                raise SecurityIdentityError(f"invalid {market} snapshot schema")
            if payload["schema_version"] != SECURITY_IDENTITY_SCHEMA_VERSION:
                raise SecurityIdentityError(f"unsupported {market} snapshot version")
            if payload["market"] != market or not isinstance(payload["records"], list):
                raise SecurityIdentityError(f"invalid {market} snapshot identity")
            market_records = tuple(SecurityRecord.from_dict(item) for item in payload["records"])
            if len(market_records) != payload["record_count"]:
                raise SecurityIdentityError(f"invalid {market} snapshot record_count")
            if any(record.market != market for record in market_records):
                raise SecurityIdentityError(f"invalid {market} snapshot record market")
            records.extend(market_records)
            loaded_markets.append(market)
            snapshots[market] = {
                "path": str(path),
                "retrieved_at": payload["retrieved_at"],
                "sources": payload["sources"],
                "record_count": payload["record_count"],
            }
        missing = tuple(item for item in requested if item not in loaded_markets)
        if require_all and missing:
            raise SecurityMasterUnavailableError(
                f"missing security-master snapshots: {', '.join(missing)}"
            )
        if not loaded_markets:
            raise SecurityMasterUnavailableError(
                f"no security-master snapshots found in {self.cache_dir}"
            )
        return SecurityMaster(
            tuple(sorted(records, key=_record_sort_key)),
            tuple(loaded_markets),
            snapshots,
        )


def load_identity_master(
    store: SecurityMasterStore,
    *,
    market: str | None = None,
) -> SecurityMaster:
    """Load enough snapshots to make the requested cross-market decision safely."""

    if market is None:
        return store.load(require_all=True)
    selected_market = _market(market)
    master = store.load()
    if selected_market not in master.markets:
        raise SecurityMasterUnavailableError(
            f"missing security-master snapshot: {selected_market}"
        )
    return master


@dataclass(frozen=True)
class IdentityCandidate:
    canonical_name: str
    market: str
    exchange: str
    ticker: str
    security_id: str
    match_basis: str
    matched_value: str
    score: float
    verified: bool
    active: bool
    source_name: str
    source_url: str
    source_record_id: str
    identifiers: dict[str, str]

    def to_dict(self) -> dict[str, Any]:
        return dict(self.__dict__)


@dataclass(frozen=True)
class IdentityResult:
    query: str
    normalized_query: str
    market_hint: str | None
    exchange_hint: str | None
    status: IdentityStatus
    reason: str
    resolved: IdentityCandidate | None
    candidates: tuple[IdentityCandidate, ...]
    schema_version: str = SECURITY_IDENTITY_SCHEMA_VERSION

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "query": self.query,
            "normalized_query": self.normalized_query,
            "market_hint": self.market_hint,
            "exchange_hint": self.exchange_hint,
            "status": self.status.value,
            "reason": self.reason,
            "resolved": self.resolved.to_dict() if self.resolved else None,
            "candidates": [item.to_dict() for item in self.candidates],
        }


def _candidate_sort_key(candidate: IdentityCandidate) -> tuple[Any, ...]:
    return (
        -candidate.score,
        _MARKET_ORDER[candidate.market],
        candidate.exchange,
        candidate.security_id,
    )


class SecurityIdentityResolver:
    """Resolve a fuzzy company query to one verified market security or stop."""

    def __init__(self, master: SecurityMaster):
        if not isinstance(master, SecurityMaster):
            raise TypeError("master must be SecurityMaster")
        self.master = master

    @staticmethod
    def _candidate(
        record: SecurityRecord,
        *,
        basis: str,
        matched_value: str,
        score: float,
    ) -> IdentityCandidate:
        return IdentityCandidate(
            canonical_name=record.canonical_name,
            market=record.market,
            exchange=record.exchange,
            ticker=record.ticker,
            security_id=record.security_id,
            match_basis=basis,
            matched_value=matched_value,
            score=round(float(score), 4),
            verified=True,
            active=record.active,
            source_name=record.source_name,
            source_url=record.source_url,
            source_record_id=record.source_record_id,
            identifiers=record.identifiers,
        )

    @staticmethod
    def _exact(record: SecurityRecord, query: str) -> IdentityCandidate | None:
        if _ticker_key(query, record.market) in {
            _ticker_key(record.ticker, record.market),
            _ticker_key(record.security_id, record.market),
        }:
            return SecurityIdentityResolver._candidate(
                record, basis="ticker_exact", matched_value=record.ticker, score=1.0
            )
        normalized = _normalize_text(query)
        if normalized == _normalize_text(record.canonical_name):
            return SecurityIdentityResolver._candidate(
                record,
                basis="official_name_exact",
                matched_value=record.canonical_name,
                score=1.0,
            )
        for alias in record.aliases:
            if normalized == _normalize_text(alias):
                return SecurityIdentityResolver._candidate(
                    record, basis="alias_exact", matched_value=alias, score=1.0
                )
        return None

    @staticmethod
    def _fuzzy(record: SecurityRecord, normalized_query: str) -> IdentityCandidate:
        values = (record.canonical_name, *record.aliases)
        scored = [
            (
                SequenceMatcher(None, normalized_query, _normalize_text(value)).ratio(),
                value,
            )
            for value in values
            if _normalize_text(value)
        ]
        score, value = max(scored, key=lambda item: (item[0], item[1]))
        return SecurityIdentityResolver._candidate(
            record, basis="strong_fuzzy", matched_value=value, score=score
        )

    @staticmethod
    def _eligible(
        candidate: IdentityCandidate,
        market: str | None,
        exchange: str | None,
    ) -> bool:
        return (market is None or candidate.market == market) and (
            exchange is None or candidate.exchange == exchange
        )

    @staticmethod
    def _result(
        query: str,
        normalized_query: str,
        market: str | None,
        exchange: str | None,
        status: IdentityStatus,
        reason: str,
        candidates: tuple[IdentityCandidate, ...],
    ) -> IdentityResult:
        resolved = candidates[0] if status is IdentityStatus.RESOLVED else None
        return IdentityResult(
            query=query,
            normalized_query=normalized_query,
            market_hint=market,
            exchange_hint=exchange,
            status=status,
            reason=reason,
            resolved=resolved,
            candidates=candidates,
        )

    def identify(
        self,
        query: str,
        *,
        market: str | None = None,
        exchange: str | None = None,
    ) -> IdentityResult:
        query = _required_text(query, "query")
        normalized_query = _normalize_text(query)
        if len(normalized_query) < 2:
            raise SecurityIdentityError("query is too short")
        market_hint = _market(market) if market else None
        exchange_hint = _required_text(exchange, "exchange").upper() if exchange else None
        exact = tuple(
            sorted(
                filter(None, (self._exact(record, query) for record in self.master.records)),
                key=_candidate_sort_key,
            )
        )
        eligible_exact = tuple(
            item
            for item in exact
            if self._eligible(item, market_hint, exchange_hint)
        )
        if len(eligible_exact) == 1:
            return self._result(
                query,
                normalized_query,
                market_hint,
                exchange_hint,
                IdentityStatus.RESOLVED,
                "one_verified_exact_identity",
                eligible_exact,
            )
        if len(eligible_exact) > 1:
            return self._result(
                query,
                normalized_query,
                market_hint,
                exchange_hint,
                IdentityStatus.AMBIGUOUS,
                "multiple_verified_exact_identities",
                eligible_exact,
            )
        if exact and (market_hint or exchange_hint):
            return self._result(
                query,
                normalized_query,
                market_hint,
                exchange_hint,
                IdentityStatus.CONFLICT,
                "exact_identity_conflicts_with_market_or_exchange_hint",
                exact,
            )

        fuzzy = tuple(
            sorted(
                (
                    self._fuzzy(record, normalized_query)
                    for record in self.master.records
                    if (market_hint is None or record.market == market_hint)
                    and (exchange_hint is None or record.exchange == exchange_hint)
                ),
                key=_candidate_sort_key,
            )
        )
        candidates = tuple(
            item for item in fuzzy if item.score >= _FUZZY_CANDIDATE_THRESHOLD
        )[:5]
        if not candidates:
            return self._result(
                query,
                normalized_query,
                market_hint,
                exchange_hint,
                IdentityStatus.MISSING,
                "no_verified_identity_candidate",
                (),
            )
        top = candidates[0]
        runner_up = candidates[1].score if len(candidates) > 1 else 0.0
        if top.score >= _FUZZY_RESOLVE_THRESHOLD and top.score - runner_up >= _FUZZY_MARGIN:
            return self._result(
                query,
                normalized_query,
                market_hint,
                exchange_hint,
                IdentityStatus.RESOLVED,
                "one_unique_strong_fuzzy_identity",
                (top,),
            )
        return self._result(
            query,
            normalized_query,
            market_hint,
            exchange_hint,
            IdentityStatus.AMBIGUOUS,
            "fuzzy_candidates_require_user_selection",
            candidates,
        )


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _cn_exchange(code: str) -> str | None:
    if code.startswith(("600", "601", "603", "605", "688", "689")):
        return "SSE"
    if code.startswith(("000", "001", "002", "003", "300", "301")):
        return "SZSE"
    if code.startswith(("4", "8", "920")):
        return "BSE"
    return None


def _parse_cninfo(payload: bytes) -> tuple[SecurityRecord, ...]:
    data = json.loads(payload.decode("utf-8-sig"))
    rows = data.get("stockList") if isinstance(data, dict) else data
    if not isinstance(rows, list):
        raise SecurityIdentityError("CNINFO stockList is missing")
    records: dict[str, SecurityRecord] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        raw_code = str(row.get("code") or "").strip()
        code = raw_code.zfill(6) if raw_code.isdigit() else raw_code
        name = str(row.get("zwjc") or "").strip()
        exchange = _cn_exchange(code)
        if not re.fullmatch(r"\d{6}", code) or not name or exchange is None:
            continue
        aliases = set(_generated_aliases(name, "CN"))
        pinyin = str(row.get("pinyin") or row.get("py") or "").strip()
        if pinyin:
            aliases.add(pinyin)
        org_id = str(row.get("orgId") or row.get("orgid") or code).strip()
        identifiers = {"org_id": org_id}
        category = str(row.get("category") or "").strip()
        if category:
            identifiers["cninfo_category"] = category
        records[code] = SecurityRecord(
            canonical_name=name,
            market="CN",
            exchange=exchange,
            ticker=code,
            security_id=code,
            aliases=tuple(sorted(aliases)),
            active=True,
            source_name="cninfo",
            source_url=CNINFO_STOCK_URL,
            source_record_id=org_id,
            identifiers=identifiers,
        )
    return tuple(sorted(records.values(), key=_record_sort_key))


def _json_rows(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        rows = [item for item in value if isinstance(item, dict)]
        if rows and any(set(item) & {"c", "stockCode", "stock_code"} for item in rows):
            return rows
        for item in value:
            nested = _json_rows(item)
            if nested:
                return nested
    elif isinstance(value, dict):
        for item in value.values():
            nested = _json_rows(item)
            if nested:
                return nested
    return []


def _first_value(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _hk_equity_codes(workbook_payload: bytes) -> dict[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SecurityIdentityError(
            "HK security-master refresh requires the catalog extra (openpyxl)"
        ) from exc
    workbook = load_workbook(io.BytesIO(workbook_payload), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers: dict[str, int] | None = None
    codes: dict[str, str] = {}
    for row in sheet.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        normalized = [_normalize_text(value) for value in values]
        if headers is None:
            try:
                headers = {
                    "code": normalized.index("stockcode"),
                    "name": normalized.index("nameofsecurities"),
                    "category": normalized.index("category"),
                    "subcategory": normalized.index("subcategory"),
                }
            except ValueError:
                continue
            continue
        code_index = headers["code"]
        name_index = headers["name"]
        category_index = headers["category"]
        subcategory_index = headers["subcategory"]
        if max(code_index, name_index, category_index, subcategory_index) >= len(values):
            continue
        raw_code = re.sub(r"\.0$", "", values[code_index])
        code = raw_code.zfill(5) if raw_code.isdigit() else raw_code
        category = values[category_index].casefold()
        subcategory = values[subcategory_index].casefold()
        if re.fullmatch(r"\d{5}", code) and category == "equity" and "equity securities" in subcategory:
            codes[code] = values[name_index]
    if headers is None:
        raise SecurityIdentityError("HKEX securities workbook header was not found")
    return codes


def _hk_standard_transfer_codes(workbook_payload: bytes) -> dict[str, str]:
    try:
        import xlrd
    except ImportError as exc:
        raise SecurityIdentityError(
            "HK security-master fallback requires the catalog extra (xlrd)"
        ) from exc
    workbook = xlrd.open_workbook(file_contents=workbook_payload, on_demand=True)
    sheet = workbook.sheet_by_index(0)
    headers: dict[str, int] | None = None
    codes: dict[str, str] = {}
    for row_index in range(sheet.nrows):
        values = [sheet.cell_value(row_index, index) for index in range(sheet.ncols)]
        normalized = [_normalize_text(str(value)) for value in values]
        if headers is None:
            try:
                headers = {
                    "code": normalized.index("stockcode"),
                    "name": normalized.index("englishstockshortname"),
                }
            except ValueError:
                continue
            continue
        raw_code_value = values[headers["code"]]
        if isinstance(raw_code_value, float) and raw_code_value.is_integer():
            raw_code = str(int(raw_code_value))
        else:
            raw_code = re.sub(r"\.0$", "", str(raw_code_value).strip())
        code = raw_code.zfill(5) if raw_code.isdigit() and len(raw_code) <= 5 else raw_code
        name = str(values[headers["name"]]).strip()
        if re.fullmatch(r"\d{5}", code) and name:
            codes[code] = name
    if headers is None:
        raise SecurityIdentityError("HKEX standard-transfer workbook header was not found")
    return codes


def _parse_hkex(
    active_payload: bytes,
    inactive_payload: bytes,
    eligible_codes: dict[str, str],
) -> tuple[SecurityRecord, ...]:
    records: dict[str, SecurityRecord] = {}
    sources = (
        (True, HKEX_ACTIVE_STOCK_URL, active_payload),
        (False, HKEX_INACTIVE_STOCK_URL, inactive_payload),
    )
    for active, source_url, payload in sources:
        rows = _json_rows(json.loads(payload.decode("utf-8-sig")))
        for row in rows:
            raw_code = _first_value(
                row, ("stockCode", "STOCK_CODE", "stock_code", "code", "CODE", "c")
            )
            code = raw_code.zfill(5) if raw_code.isdigit() and len(raw_code) <= 5 else raw_code
            if code not in eligible_codes or code in records:
                continue
            name = _first_value(
                row,
                ("companyName", "COMPANY_NAME", "company_name", "name", "NAME", "n"),
            )
            stock_id = _first_value(
                row, ("stockId", "STOCK_ID", "stock_id", "id", "ID", "s")
            )
            if not name:
                continue
            aliases = set(_generated_aliases(name, "HK"))
            english_name = eligible_codes.get(code, "")
            if english_name:
                aliases.add(english_name)
                aliases.update(_generated_aliases(english_name, "HK"))
            records[code] = SecurityRecord(
                canonical_name=name,
                market="HK",
                exchange="HKEX",
                ticker=code,
                security_id=code,
                aliases=tuple(sorted(aliases)),
                active=active,
                source_name="hkex",
                source_url=source_url,
                source_record_id=stock_id or code,
                identifiers={"hkex_stock_id": stock_id} if stock_id else {},
            )
    return tuple(sorted(records.values(), key=_record_sort_key))


def _symbol_directory(payload: bytes, *, nasdaq: bool) -> dict[str, tuple[str, str]]:
    text = payload.decode("utf-8-sig", errors="strict")
    rows = csv.DictReader(io.StringIO(text), delimiter="|")
    result: dict[str, tuple[str, str]] = {}
    exchange_map = {
        "A": "NYSE AMERICAN",
        "N": "NYSE",
        "P": "NYSE ARCA",
        "Z": "BATS",
        "V": "IEX",
    }
    for row in rows:
        symbol_field = "Symbol" if nasdaq else "ACT Symbol"
        symbol = str(row.get(symbol_field) or "").strip().upper()
        if not symbol or symbol.startswith("FILE CREATION TIME"):
            continue
        if str(row.get("Test Issue") or "").strip().upper() == "Y":
            continue
        if str(row.get("ETF") or "").strip().upper() == "Y":
            continue
        exchange = "NASDAQ" if nasdaq else exchange_map.get(
            str(row.get("Exchange") or "").strip().upper()
        )
        if exchange is None:
            continue
        security_name = str(row.get("Security Name") or "").strip()
        result[_ticker_key(symbol, "US")] = (exchange, security_name)
    return result


def _parse_us(sec_payload: bytes, nasdaq_payload: bytes, other_payload: bytes) -> tuple[SecurityRecord, ...]:
    sec_data = json.loads(sec_payload.decode("utf-8-sig"))
    if not isinstance(sec_data, dict):
        raise SecurityIdentityError("SEC company_tickers payload must be an object")
    directory = {
        **_symbol_directory(other_payload, nasdaq=False),
        **_symbol_directory(nasdaq_payload, nasdaq=True),
    }
    records: dict[str, SecurityRecord] = {}
    for row in sec_data.values():
        if not isinstance(row, dict):
            continue
        ticker = str(row.get("ticker") or "").strip().upper()
        title = str(row.get("title") or "").strip()
        key = _ticker_key(ticker, "US")
        if not key or not title or key not in directory:
            continue
        exchange, directory_name = directory[key]
        cik_raw = row.get("cik_str")
        try:
            cik = f"{int(cik_raw):010d}"
        except (TypeError, ValueError):
            continue
        aliases = set(_generated_aliases(title, "US"))
        if directory_name:
            aliases.add(directory_name)
            aliases.update(_generated_aliases(directory_name, "US"))
        records[ticker] = SecurityRecord(
            canonical_name=title,
            market="US",
            exchange=exchange,
            ticker=ticker,
            security_id=ticker,
            aliases=tuple(sorted(aliases)),
            active=True,
            source_name="sec_nasdaq",
            source_url=SEC_TICKER_URL,
            source_record_id=cik,
            identifiers={"cik": cik},
        )
    return tuple(sorted(records.values(), key=_record_sort_key))


class OfficialSecurityMasterRefresher:
    """Refresh each official market snapshot independently and keep stale successes."""

    def __init__(
        self,
        store: SecurityMasterStore,
        *,
        fetch_bytes: Callable[[str], bytes] | None = None,
        clock: Callable[[], str] = _utc_now,
        minimum_records: dict[str, int] | None = None,
    ):
        if not isinstance(store, SecurityMasterStore):
            raise TypeError("store must be SecurityMasterStore")
        self.store = store
        self.fetch_bytes = fetch_bytes or self._fetch
        self.clock = clock
        self.minimum_records = dict(_DEFAULT_MINIMUM_RECORDS)
        for raw_market, raw_value in (minimum_records or {}).items():
            market = _market(raw_market)
            if isinstance(raw_value, bool) or not isinstance(raw_value, int) or raw_value < 1:
                raise ValueError(f"minimum_records[{market}] must be a positive integer")
            self.minimum_records[market] = raw_value

    @staticmethod
    def _fetch(url: str) -> bytes:
        user_agent = os.environ.get(
            "SEC_USER_AGENT", "company-wiki/0.1 source-catalog security-master"
        )
        response = requests.get(
            url,
            headers={"User-Agent": user_agent, "Accept": "*/*"},
            timeout=30,
        )
        response.raise_for_status()
        return response.content

    def _refresh_market(self, market: str) -> tuple[tuple[SecurityRecord, ...], tuple[str, ...]]:
        if market == "CN":
            return _parse_cninfo(self.fetch_bytes(CNINFO_STOCK_URL)), (CNINFO_STOCK_URL,)
        if market == "HK":
            active_payload = self.fetch_bytes(HKEX_ACTIVE_STOCK_URL)
            inactive_payload = self.fetch_bytes(HKEX_INACTIVE_STOCK_URL)
            primary_codes = _hk_equity_codes(self.fetch_bytes(HKEX_SECURITIES_URL))
            records = _parse_hkex(active_payload, inactive_payload, primary_codes)
            sources = [HKEX_ACTIVE_STOCK_URL, HKEX_INACTIVE_STOCK_URL, HKEX_SECURITIES_URL]
            if len(records) < self.minimum_records["HK"]:
                fallback_codes = _hk_standard_transfer_codes(
                    self.fetch_bytes(HKEX_STANDARD_TRANSFER_URL)
                )
                fallback_records = _parse_hkex(active_payload, inactive_payload, fallback_codes)
                if len(fallback_records) > len(records):
                    records = fallback_records
                sources.append(HKEX_STANDARD_TRANSFER_URL)
            return (
                records,
                tuple(sources),
            )
        if market == "US":
            return (
                _parse_us(
                    self.fetch_bytes(SEC_TICKER_URL),
                    self.fetch_bytes(NASDAQ_LISTED_URL),
                    self.fetch_bytes(NASDAQ_OTHER_LISTED_URL),
                ),
                (SEC_TICKER_URL, NASDAQ_LISTED_URL, NASDAQ_OTHER_LISTED_URL),
            )
        raise SecurityIdentityError(f"unsupported market: {market}")

    def refresh(self, *, markets: tuple[str, ...] = SECURITY_MARKETS) -> dict[str, Any]:
        selected = tuple(dict.fromkeys(_market(item) for item in markets))
        report: dict[str, Any] = {}
        for market in selected:
            retrieved_at = self.clock()
            try:
                records, sources = self._refresh_market(market)
                minimum = self.minimum_records[market]
                if len(records) < minimum:
                    raise SecurityIdentityError(
                        f"{market} refresh returned {len(records)} company securities; minimum is {minimum}"
                    )
                path = self.store.write_market(
                    market,
                    records,
                    retrieved_at=retrieved_at,
                    sources=sources,
                )
                report[market] = {
                    "status": "refreshed",
                    "records": len(records),
                    "retrieved_at": retrieved_at,
                    "path": str(path),
                    "sources": list(sources),
                    "error": None,
                }
            except Exception as exc:
                report[market] = {
                    "status": "stale_cache" if self.store.has_market(market) else "failed",
                    "records": None,
                    "retrieved_at": None,
                    "path": str(self.store.path_for(market)),
                    "sources": None,
                    "error": f"{type(exc).__name__}: {str(exc)[:500]}",
                }
        return report


__all__ = [
    "CNINFO_STOCK_URL",
    "HKEX_ACTIVE_STOCK_URL",
    "HKEX_INACTIVE_STOCK_URL",
    "HKEX_SECURITIES_URL",
    "HKEX_STANDARD_TRANSFER_URL",
    "IdentityCandidate",
    "IdentityResult",
    "IdentityStatus",
    "NASDAQ_LISTED_URL",
    "NASDAQ_OTHER_LISTED_URL",
    "OfficialSecurityMasterRefresher",
    "SECURITY_IDENTITY_SCHEMA_VERSION",
    "SECURITY_MARKETS",
    "SEC_TICKER_URL",
    "SecurityIdentityError",
    "SecurityIdentityResolutionError",
    "SecurityIdentityResolver",
    "SecurityMaster",
    "SecurityMasterStore",
    "SecurityMasterUnavailableError",
    "SecurityRecord",
    "load_identity_master",
]
